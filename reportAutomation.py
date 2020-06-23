# Aouthor Beomsang Kim
# Produced 21-Jun-2020
# Origin netScout
# Purpose   지난주와 이번주 리포트의 날짜를 입력받음
#           금주의 직워 리포트 취합
#           취합된 금주의 리포트들을 양식에 맞춰 새로운 하나의 리포트로 만듬
# Comment   input은 다른 파일에서 따로 받을려 했는데 실패함


import openpyxl as xl
import os

lastWeekDate = ""
thisWeekDate = ""

print("양식 비교를 위함입니다")
print("지난주 리포트 생성 날자를 입력해 주세요")
lastWeekDate = input("YYYYMMDD: ")
print("이번주 리포트 생성 날자를 입력해 주세요")
thisWeekDate = input("YYYYMMDD: ")
print("잠시만 기다려 주세요")

#c:\sources
currentPath = os.path.dirname(__file__)
#c:\sources + \리포트 =>
reportPath = os.path.join(currentPath, "리포트")

reports = []

def read():
    for file in os.listdir(reportPath):
        if file.endswith(".xlsx") and thisWeekDate in file:
            filePath = os.path.join(reportPath, file)
            
            wb = xl.load_workbook(filePath)
            sheet = wb.active

            #3행부터 1열-5열
            #name, lastweek, thisweek, nextweek, issue
            row = 3
            col = 1

            #10번 반복하겠다(팀원이 10명을 넘지 않는다는 가정)
            for _ in range(10):
                name = sheet.cell(row = row, column = col).value
                if name is None:
                    break
                lastweek = sheet.cell(row = row, column = col+1).value
                thisweek = sheet.cell(row = row, column = col+2).value
                nextweek = sheet.cell(row = row, column = col+3).value
                issue = sheet.cell(row = row, column = col+4).value

                reports.append({"name":name, "lastweek":lastweek, "thisweek":thisweek, "nextweek":nextweek, "issue":issue})
                row += 1
    return reports
            

lastWeekReport = read()

lastWeekReportTitle = os.path.join(reportPath, "Weekly Report_"+ lastWeekDate +".xlsx")

#지난주 리포트를 읽어오고
wb = xl.load_workbook(lastWeekReportTitle)
sheet = wb.active

nameList = {} #[] {}
newRow = 0

thisreport = {}

def readAndWrite():

    for i in range(3, 13):
        name = sheet.cell(row = i, column = 1).value
        if name is None:
            newRow = i
            break
        #key : value
        #"홍길동" : 3
        #"영희" : 4
        nameList[name] = i

    #지난 주 리포트에 저장되어 있는 팀원들의 위치 정보를 확인

    #읽어온 데이터를 바탕으로 새로 데이터를 작성한다
    col = 1
    for r in lastWeekReport:
        row = nameList.get(r["name"], -1)
        if row == -1:
            row = newRow
            newRow += 1
        
        sheet.cell(row = row, column = col).value = r["name"]
        sheet.cell(row = row, column = col+1).value = r["lastweek"]
        sheet.cell(row = row, column = col+2).value = r["thisweek"]
        sheet.cell(row = row, column = col+3).value = r["nextweek"]
        sheet.cell(row = row, column = col+4).value = r["issue"]


    thisreport = wb.save(os.path.join(reportPath, "Weekly Report_"+ thisWeekDate +".xlsx"))

    return thisreport

newReport = readAndWrite()

print(thisWeekDate + " 일자 리포트 생성이 완료되었습니다")


