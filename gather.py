# Aouthor Beomsang Kim
# Produced 21-Jun-2020
# Origin netScout
# Purpose 금주의 레포트 들을 취합


import openpyxl as xl
import os
import reportAutomation



#c:\sources
currentPath = os.path.dirname(__file__)
#c:\sources + \리포트 =>
reportPath = os.path.join(currentPath, "리포트")

reports = []

def gathering() :
    for file in os.listdir(reportPath):
        if file.endswith(".xlsx") and reportAutomation.thisWeekDate in file:
            filePath = os.path.join(reportPath, file)
            
            wb = xl.load_workbook(filePath)
            sheet = wb.active

            #3행부터 1열-5열
            #name, lastweek, thisweek, nextweek, issue
            row = 3
            col = 1

            #10번 반복하겠다(팀원이 10명을 넘지 않는다는 가정)
            for _ in range(10):
                name = sheet.cell(row = row, column = col).value
                if name is None:
                    break
                lastweek = sheet.cell(row = row, column = col+1).value
                thisweek = sheet.cell(row = row, column = col+2).value
                nextweek = sheet.cell(row = row, column = col+3).value
                issue = sheet.cell(row = row, column = col+4).value

                reports.append({"name":name, "lastweek":lastweek, "thisweek":thisweek, "nextweek":nextweek, "issue":issue})
                row += 1
        return reports


