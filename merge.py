# Aouthor Beomsang Kim
# Produced 21-Jun-2020
# Origin netScout
# Purpose 취합된 금주의 리포트들을 양식에 맞춰 새로운 하나의 리포트로 만듬

import gather
import reportAutomation
import openpyxl as xl
import os


#지난주 리포트를 읽어오고
wb = xl.load_workbook(os.path.join(gather.reportPath, "Weekly Report_"+ reportAutomation.lastWeekDate +".xlsx"))
sheet = wb.active

nameList = {} #[] {}
newRow = 0
newReport = gather.gathering()

def merging():

    for i in range(3, 13):
        name = sheet.cell(row = i, column = 1).value
        if name is None:
            newRow = i
            break
        #key : value
        #"홍길동" : 3
        #"영희" : 4
        nameList[name] = i

    #지난 주 리포트에 저장되어 있는 팀원들의 위치 정보를 확인

    #읽어온 데이터를 바탕으로 새로 데이터를 작성한다
    col = 1
    for r in newReport:
        row = nameList.get(r["name"], -1)
        if row == -1:
            row = newRow
            newRow += 1
        
        sheet.cell(row = row, column = col).value = r["name"]
        sheet.cell(row = row, column = col+1).value = r["lastweek"]
        sheet.cell(row = row, column = col+2).value = r["thisweek"]
        sheet.cell(row = row, column = col+3).value = r["nextweek"]
        sheet.cell(row = row, column = col+4).value = r["issue"]

title = "Weekly Report_"+ reportAutomation.thisWeekDate +".xlsx"
wb.save(os.path.join(gather.reportPath, title))